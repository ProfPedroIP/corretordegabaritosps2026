import streamlit as st
import cv2
import numpy as np
import pandas as pd
from pdf2image import convert_from_path, pdfinfo_from_path 
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import tempfile
import io
import gc 

# Configuração da página Web
st.set_page_config(page_title="Corretor de gabaritos", page_icon="📝", layout="centered")

# --- EXIBIÇÃO DA LOGO CENTRALIZADA ---
URL_LOGO = "https://www.institutoponte.org.br/wp-content/uploads/2025/02/Logo-Instituto-Ponto.png"
col1, col2, col3 = st.columns([1, 1, 1])
with col2:
    st.image(URL_LOGO, width=250)

def isolar_blocos_com_protecao(imagem_cv):
    altura_total, largura_total = imagem_cv.shape[:2]
    y_limite_superior = int(altura_total * 0.30)
    area_min, area_max = 1100, 1400 
    
    cinza = cv2.cvtColor(imagem_cv, cv2.COLOR_BGR2GRAY)
    desfoque = cv2.GaussianBlur(cinza, (5, 5), 0)
    _, binario = cv2.threshold(desfoque, 150, 255, cv2.THRESH_BINARY_INV)
    contornos, _ = cv2.findContours(binario, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    marcadores = []
    
    for c in contornos:
        perimetro = cv2.arcLength(c, True)
        aprox = cv2.approxPolyDP(c, 0.05 * perimetro, True)
        area = cv2.contourArea(c)
        if len(aprox) == 4 and area_min <= area <= area_max:
            x, y, w, h = cv2.boundingRect(aprox)
            if 0.7 <= (w / float(h)) <= 1.3:
                cx, cy = x + (w//2), y + (h//2)
                if cy > y_limite_superior:
                    marcadores.append([cx, cy])

    if len(marcadores) < 8: return None, None
    x_meio = largura_total / 2
    m_esq = [p for p in marcadores if p[0] < x_meio]
    m_dir = [p for p in marcadores if p[0] >= x_meio]
    if len(m_esq) < 4 or len(m_dir) < 4: return None, None

    def ordenar_4(pts):
        pts_arr = np.array(pts)
        ret = np.zeros((4, 2), dtype="float32")
        s = pts_arr.sum(axis=1)
        ret[0], ret[2] = pts_arr[np.argmin(s)], pts_arr[np.argmax(s)]
        d = np.diff(pts_arr, axis=1)
        ret[1], ret[3] = pts_arr[np.argmin(d)], pts_arr[np.argmax(d)]
        return ret

    def processar(pts):
        origem = ordenar_4(pts)
        destino = np.array([[0,0], [628,0], [628,1093], [0,1093]], dtype="float32")
        matriz = cv2.getPerspectiveTransform(origem, destino)
        return cv2.warpPerspective(imagem_cv, matriz, (629, 1094))

    return processar(m_esq), processar(m_dir)

def ler_bolinhas(img_bloco, q_ini):
    cinza = cv2.cvtColor(img_bloco, cv2.COLOR_BGR2GRAY)
    _, binario = cv2.threshold(cinza, 210, 255, cv2.THRESH_BINARY_INV)
    respostas = {}
    alts = ['A', 'B', 'C', 'D', 'E']
    xi, yi, px, py, raio, limite = 89, 78, 110, 104, 31, 0.30

    for i in range(10): 
        marcadas = []
        for j in range(5): 
            cx, cy = xi + (j * px), yi + (i * py)
            celula = binario[cy-raio : cy+raio, cx-raio : cx+raio]
            if celula.size > 0 and (cv2.countNonZero(celula) / celula.size) > limite:
                marcadas.append(alts[j])
        
        if len(marcadas) == 0: respostas[q_ini+i] = "EM BRANCO"
        elif len(marcadas) > 1: respostas[q_ini+i] = "ANULADA"
        else: respostas[q_ini+i] = marcadas[0]
    return respostas

# --- INTERFACE VISUAL DO SITE ---
st.title("Correção Automática de Gabaritos")
st.markdown("Sistema de correção oficial do Processo Seletivo 2026 do **Instituto Ponte**")

with st.expander("📖 Leia as Instruções Importantes", expanded=True):
    st.write("1. Digitalize os gabaritos utilizando o **Adobe Scan** e gere um arquivo PDF.")
    st.write("2. **Importante:** Este programa não lê nomes. Os resultados seguem a ordem exata das páginas do PDF.")
    st.write("3. **Dica:** Enumere fisicamente os gabaritos antes de escanear.")

st.subheader("1. Identificação e Gabarito")
c1, c2 = st.columns(2)
serie_escolhida = c1.selectbox("Selecione a Série:", ["7º Ano", "8º Ano", "9º Ano", "1ª Série", "2ª Série"])
polo_escolhido = c2.text_input("Polo:", placeholder="Ex: Bela Cruz")

st.write("Preencha o Gabarito Correto para essa prova:")
padrao = "A B C D E A B C D E A B C D E A B C D E".split()
gabarito_inputs = {}

st.markdown("#### 📚 Português (Questões 01 a 10)")
cols = st.columns(10)
for i in range(1, 11):
    gabarito_inputs[i] = cols[i-1].text_input(f"Q{i}", padrao[i-1], key=f"q{i}", max_chars=1).upper()

st.markdown("#### 📐 Matemática (Questões 11 a 20)")
cols2 = st.columns(10)
for i in range(11, 21):
    gabarito_inputs[i] = cols2[i-11].text_input(f"Q{i}", padrao[i-1], key=f"q{i}", max_chars=1).upper()

st.divider()

st.subheader("2. Envio de Arquivos")
arquivos_pdf = st.file_uploader("Arraste aqui o(s) PDF(s) digitalizado(s)", type=["pdf"], accept_multiple_files=True)

if st.button("🚀 Executar Correção dos Gabaritos", type="primary"):
    if not arquivos_pdf:
        st.warning("Por favor, faça o upload de pelo menos um arquivo PDF.")
    elif not polo_escolhido:
        st.error("Por favor, preencha o campo 'Polo'.")
    else:
        # --- NOVO: PRÉ-CONTAGEM DE TODAS AS PÁGINAS PARA A BARRA DE PROGRESSO ---
        total_geral_paginas = 0
        arquivos_processados = [] # Guardar infos para não repetir leitura de disco
        
        with st.spinner("Calculando volume de trabalho..."):
            for file in arquivos_pdf:
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                    tmp.write(file.read())
                    tmp_path = tmp.name
                    info = pdfinfo_from_path(tmp_path)
                    total_geral_paginas += info["Pages"]
                    arquivos_processados.append((tmp_path, file.name, info["Pages"]))

        dados_consolidados = []
        num_global = 1
        progress_bar = st.progress(0)
        status_text = st.empty()

        # --- LOOP DE CORREÇÃO REAL ---
        for tmp_path, nome_original, total_pags_arquivo in arquivos_processados:
            for p in range(1, total_pags_arquivo + 1):
                # Atualização da barra: (Gabaritos já feitos) / (Total de todos os arquivos)
                progresso_atual = (num_global - 1) / total_geral_paginas
                progress_bar.progress(progresso_atual)
                
                status_text.text(f"Corrigindo Gabarito Nº {num_global:04d} de {total_geral_paginas}...")
                
                # Processamento a 300 DPI página por página
                pagina_imagem = convert_from_path(tmp_path, dpi=300, first_page=p, last_page=p)[0]
                img = cv2.cvtColor(np.array(pagina_imagem), cv2.COLOR_RGB2BGR)
                del pagina_imagem 
                
                bloco_e, bloco_d = isolar_blocos_com_protecao(img)
                resp = {}
                
                if bloco_e is not None and bloco_d is not None:
                    resp.update(ler_bolinhas(bloco_e, 1))
                    resp.update(ler_bolinhas(bloco_d, 11))
                    acertos_pt = sum(1 for q in range(1, 11) if resp.get(q) == gabarito_inputs[q])
                    acertos_mt = sum(1 for q in range(11, 21) if resp.get(q) == gabarito_inputs[q])
                else:
                    acertos_pt, acertos_mt = 0, 0
                    for q in range(1, 21): resp[q] = "ERRO_LEITURA"

                linha = {"Questão/Gabarito": f"Nº {num_global:04d}"}
                for q in range(1, 21): linha[f"Q{q}"] = resp.get(q)
                linha["Português"] = acertos_pt
                linha["Matemática"] = acertos_mt
                dados_consolidados.append(linha)
                
                num_global += 1
                gc.collect()

        # Finaliza a barra em 100%
        progress_bar.progress(1.0)

        if dados_consolidados:
            df = pd.DataFrame(dados_consolidados)
            output = io.BytesIO()
            df.to_excel(output, index=False)
            output.seek(0)
            
            wb = load_workbook(output)
            ws = wb.active

            # Linha do Gabarito Esperado
            ws.insert_rows(2)
            ws.cell(row=2, column=1).value = "Gabarito Correto"
            ws.cell(row=2, column=1).font = Font(bold=True)
            for q in range(1, 21):
                cell = ws.cell(row=2, column=q+1)
                cell.value = gabarito_inputs[q]
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(bold=True)

            # Pintura de Células
            COR_ACERTO = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            COR_ERRO = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            COR_ANULADA = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

            for row in range(3, ws.max_row + 1):
                for col in range(2, 22):
                    celula = ws.cell(row=row, column=col)
                    valor = celula.value
                    questao_num = col - 1
                    esperado = gabarito_inputs[questao_num]
                    if valor == esperado: celula.fill = COR_ACERTO
                    elif valor == "ANULADA": celula.fill = COR_ANULADA
                    elif valor not in ["EM BRANCO", "ERRO_LEITURA"]: celula.fill = COR_ERRO

            final_output = io.BytesIO()
            wb.save(final_output)
            final_output.seek(0)

            nome_arquivo_final = f"Resultados - {serie_escolhida} - {polo_escolhido}.xlsx"
            status_text.empty()
            st.success(f"✅ Concluído! O Super Perseu corrigiu {len(dados_consolidados)} gabaritos com sucesso.")
            st.download_button(
                label="📥 Baixar Planilha de Resultados",
                data=final_output,
                file_name=nome_arquivo_final,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
